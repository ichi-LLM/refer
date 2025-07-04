"""
Excel ファイルの読み書きを処理
要件一覧とDescription編集の2シート構造を管理
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import logging
from typing import List, Dict, Optional, Tuple
from html.parser import HTMLParser
import re
import os

logger = logging.getLogger(__name__)


class HTMLTableParser(HTMLParser):
    """HTMLテーブルをパースするクラス"""
    
    def __init__(self):
        super().__init__()
        self.tables = []
        self.current_table = []
        self.current_row = []
        self.current_cell = ""
        self.in_table = False
        self.in_row = False
        self.in_cell = False
        
    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self.in_table = True
            self.current_table = []
        elif tag == "tr" and self.in_table:
            self.in_row = True
            self.current_row = []
        elif tag in ["td", "th"] and self.in_row:
            self.in_cell = True
            self.current_cell = ""
            
    def handle_endtag(self, tag):
        if tag == "table":
            self.in_table = False
            if self.current_table:
                self.tables.append(self.current_table)
        elif tag == "tr" and self.in_row:
            self.in_row = False
            if self.current_row:
                self.current_table.append(self.current_row)
        elif tag in ["td", "th"] and self.in_cell:
            self.in_cell = False
            self.current_row.append(self.current_cell.strip())
            
    def handle_data(self, data):
        if self.in_cell:
            self.current_cell += data


class ExcelHandler:
    """Excelファイルの読み書きハンドラー"""
    
    # カラム定義
    COLUMNS = {
        'A': 'JAMA_ID',
        'B': 'メモ/コメント',
        'C': 'Sequence',
        'D': '階層1',
        'E': '階層2',
        'F': '階層3',
        'G': '階層4',
        'H': '階層5',
        'I': '階層6',
        'J': '階層7',
        'K': '階層8',
        'L': '階層9',
        'M': '階層10',
        'N': '階層11',
        'O': 'アイテムタイプ',
        'P': 'Assignee',
        'Q': 'Status',
        'R': 'Tags',
        'S': 'Reason',
        'T': 'Preconditions',
        'U': 'Target_system',
        'V': '現在のDescription',
        'W': '要件更新',
        'X': '新Description参照'
    }
    
    # System Description編集シートの列幅（拡張性を持たせる）
    SYSTEM_DESC_COLS = {
        'a': 1,      # (a)Trigger action: 1列
        'b': 67,     # (b)Behavior of ego-vehicle: 67列（64から拡張）
        'c': 10,     # (c)HMI: 10列
        'd': 5       # (d)Other: 5列
    }
    
    # User Description編集シートの列幅（拡張性を持たせる）
    USER_DESC_COLS = {
        'a': 1,      # (a)Trigger action: 1列
        'c': 6       # (c)HMI: 6列
    }
    
    # 固定列幅設定
    FIXED_COLUMN_WIDTH_A = 16  # A列の幅
    FIXED_COLUMN_WIDTH_OTHER = 24  # その他の列の幅
    
    # デフォルトフォント設定
    DEFAULT_FONT_NAME = 'BIZ UDPゴシック'
    
    # 新規System Requirement用テンプレート数
    NEW_SYSTEM_REQUIREMENT_TEMPLATE_COUNT = 200
    
    # 新規User Requirement用テンプレート数
    NEW_USER_REQUIREMENT_TEMPLATE_COUNT = 200
    
    def __init__(self, config=None):
        """初期化"""
        self.wb = None
        self.requirement_sheet = None
        self.system_description_sheet = None
        self.user_description_sheet = None
        self.sequence_index = {}  # sequence検索用インデックス
        self.system_requirement_template_map = {}  # System Requirementテンプレートのマッピング
        self.user_requirement_template_map = {}  # User Requirementテンプレートのマッピング
        self.all_existing_items = []  # バリデーション用の既存アイテム
        
        # 設定オブジェクトを保持
        self.config = config
        
        # パフォーマンス設定
        if config:
            self.column_width_check_rows = config.column_width_check_rows
        else:
            self.column_width_check_rows = 100
            
    @property
    def system_desc_total_cols(self):
        """System Description編集シートの総列数を動的に計算"""
        # 1(行ヘッダー) + 各項目の列数の合計
        return 1 + sum(self.SYSTEM_DESC_COLS.values())
        
    @property
    def user_desc_total_cols(self):
        """User Description編集シートの総列数を動的に計算"""
        # 1(行ヘッダー) + 各項目の列数の合計
        return 1 + sum(self.USER_DESC_COLS.values())
        
    def _get_system_column_positions(self):
        """System Description用の各項目の列位置を動的に計算"""
        positions = {}
        col_idx = 2  # B列から開始（A列は行ヘッダー）
        
        # (a) Trigger action
        positions['a_start'] = col_idx
        positions['a_end'] = col_idx + self.SYSTEM_DESC_COLS['a'] - 1
        col_idx = positions['a_end'] + 1
        
        # (b) Behavior of ego-vehicle
        positions['b_start'] = col_idx
        positions['b_end'] = col_idx + self.SYSTEM_DESC_COLS['b'] - 1
        col_idx = positions['b_end'] + 1
        
        # (c) HMI
        positions['c_start'] = col_idx
        positions['c_end'] = col_idx + self.SYSTEM_DESC_COLS['c'] - 1
        col_idx = positions['c_end'] + 1
        
        # (d) Other
        positions['d_start'] = col_idx
        positions['d_end'] = col_idx + self.SYSTEM_DESC_COLS['d'] - 1
        
        return positions
        
    def _get_user_column_positions(self):
        """User Description用の各項目の列位置を動的に計算"""
        positions = {}
        col_idx = 2  # B列から開始（A列は行ヘッダー）
        
        # (a) Trigger action
        positions['a_start'] = col_idx
        positions['a_end'] = col_idx + self.USER_DESC_COLS['a'] - 1
        col_idx = positions['a_end'] + 1
        
        # (c) HMI
        positions['c_start'] = col_idx
        positions['c_end'] = col_idx + self.USER_DESC_COLS['c'] - 1
        
        return positions
        
    def create_requirement_excel(self, items: List[Dict], output_file: str, show_progress: bool = True) -> None:
        """
        要件一覧をExcelファイルに出力
        
        Args:
            items: 要件アイテムリスト
            output_file: 出力ファイル名
            show_progress: 進捗表示フラグ
        """
        logger.info(f"Excelファイル作成開始: {output_file}")
        logger.info(f"処理対象アイテム数: {len(items)}件")
        
        # 新規ワークブック作成
        logger.info("新規ワークブック作成中...")
        self.wb = Workbook()
        
        # デフォルトフォントを設定
        self.wb._default_font = Font(name=self.DEFAULT_FONT_NAME)
        
        # シート作成順序を変更：User → System → Requirement
        
        # 1. User_Description_editシートを作成
        if show_progress:
            print("\n[2/4] User Description編集シート作成中...")
        logger.info("User Description編集シート作成開始...")
        self.user_description_sheet = self.wb.create_sheet("User_Description_edit")
        self._create_user_description_sheet(items)
        
        # 2. System_Description_editシートを作成
        if show_progress:
            print("\n[2/4] System Description編集シート作成中...")
        logger.info("System Description編集シート作成開始...")
        self.system_description_sheet = self.wb.create_sheet("System_Description_edit")
        self._create_system_description_sheet(items)
        
        # 3. Requirement_of_Driverシートを作成
        if show_progress:
            print("\n[3/4] 要件一覧シート作成中...")
        logger.info("要件一覧シート作成開始...")
        self.requirement_sheet = self.wb.active
        self.requirement_sheet.title = "Requirement_of_Driver"
        self._create_requirement_sheet(items)
        
        # 保存
        logger.info("ファイル保存中...")
        self.wb.save(output_file)
        logger.info(f"Excelファイル作成完了: {output_file}")
        
    def _create_requirement_sheet(self, items: List[Dict]) -> None:
        """要件一覧シートを作成"""
        ws = self.requirement_sheet
        
        # ヘッダー行作成
        logger.info("ヘッダー行作成中...")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name=self.DEFAULT_FONT_NAME, color="FFFFFF", bold=True)
        
        for col, header in self.COLUMNS.items():
            cell = ws[f"{col}1"]
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # sequence検索用インデックスを作成（高速化のため）
        logger.info("階層情報インデックス作成中...")
        self.sequence_index = {item.get("sequence", ""): item for item in items if item.get("sequence")}
        logger.info(f"インデックス作成完了: {len(self.sequence_index)}件")
            
        # データ行作成
        total_items = len(items)
        logger.info(f"データ行作成開始: {total_items}件")
        
        # System/User Requirementカウンター初期化
        self._system_requirement_count = 0
        self._user_requirement_count = 0
        
        row_num = 2
        for idx, item in enumerate(items, 1):
            # 進捗表示（100件ごと、または最後）
            if idx % 100 == 0 or idx == total_items:
                logger.info(f"要件シート作成進捗: {idx}/{total_items} ({idx/total_items*100:.1f}%)")
                
            # 階層情報を解析（高速化版）
            hierarchy = self._parse_hierarchy_fast(item)
            
            # 基本情報（フォント設定付き）
            cell = ws[f"A{row_num}"]
            cell.value = item.get("jama_id", "")
            cell.font = Font(name=self.DEFAULT_FONT_NAME)
            
            cell = ws[f"B{row_num}"]
            cell.value = ""  # メモ/コメント欄（空欄）
            cell.font = Font(name=self.DEFAULT_FONT_NAME)
            
            cell = ws[f"C{row_num}"]
            cell.value = item.get("sequence", "")
            cell.font = Font(name=self.DEFAULT_FONT_NAME)
            
            # 階層1～11
            for i, level in enumerate(hierarchy, 1):
                if i <= 11:
                    cell = ws[f"{get_column_letter(3 + i)}{row_num}"]
                    cell.value = level
                    cell.font = Font(name=self.DEFAULT_FONT_NAME)
                    
            # その他の情報
            for col, field, default in [
                ("O", "item_type", "Requirement"),
                ("P", "assignee", ""),
                ("Q", "status", ""),
                ("R", "tags", ""),
                ("S", "reason", ""),
                ("T", "preconditions", ""),
                ("U", "target_system", "")
            ]:
                cell = ws[f"{col}{row_num}"]
                if field == "item_type":
                    # アイテムタイプを実際のIDから名前に変換
                    if item.get("item_type_id"):
                        # 既存要件：実際のIDから名前を取得
                        type_name = self.config.get_item_type_name(item["item_type_id"]) if self.config else "Unknown"
                        cell.value = type_name
                    else:
                        # 新規要件：階層と名前から推定
                        # 階層情報から階層レベルを計算
                        hierarchy_level = len(hierarchy)
                        if hierarchy_level in [10, 11]:
                            # 10-11階層は要件名で判定
                            item_type_id = self._determine_item_type_by_name_for_excel(item.get("name", ""), hierarchy_level)
                        else:
                            # 1-9階層は設定から取得
                            item_type_id = self.config.get_item_type_for_level(hierarchy_level) if self.config else None
                        
                        if item_type_id and self.config:
                            cell.value = self.config.get_item_type_name(item_type_id)
                        else:
                            cell.value = default
                else:
                    cell.value = item.get(field, default)
                cell.font = Font(name=self.DEFAULT_FONT_NAME)
            
            # Description関連
            description = item.get("description", "")
            if description:
                # HTMLテーブルを簡易表示
                cell = ws[f"V{row_num}"]
                cell.value = self._extract_table_preview(description)
                cell.font = Font(name=self.DEFAULT_FONT_NAME)
                
                # System RequirementまたはUser RequirementのDescriptionがある場合は編集リンクを作成
                item_type_id = item.get("item_type_id")
                if item_type_id in [301, 266]:
                    cell = ws[f"W{row_num}"]
                    cell.value = ""  # デフォルトは空欄（更新しない）
                    cell.font = Font(name=self.DEFAULT_FONT_NAME)
                    
                    # 適切なテンプレートマップから正しいテンプレート位置を取得
                    if item_type_id == 301 and hasattr(self, 'system_requirement_template_map') and (idx - 1) in self.system_requirement_template_map:
                        template_row = self.system_requirement_template_map[idx - 1]
                        cell = ws[f"X{row_num}"]
                        cell.value = "編集画面へ"
                        # 要件名が見えるように、1行上にリンク
                        cell.hyperlink = f"#System_Description_edit!A{template_row - 1}"
                        cell.font = Font(name=self.DEFAULT_FONT_NAME, color="0000FF", underline="single")
                    elif item_type_id == 266 and hasattr(self, 'user_requirement_template_map') and (idx - 1) in self.user_requirement_template_map:
                        template_row = self.user_requirement_template_map[idx - 1]
                        cell = ws[f"X{row_num}"]
                        cell.value = "編集画面へ"
                        # 要件名が見えるように、1行上にリンク
                        cell.hyperlink = f"#User_Description_edit!A{template_row - 1}"
                        cell.font = Font(name=self.DEFAULT_FONT_NAME, color="0000FF", underline="single")
                else:
                    cell = ws[f"W{row_num}"]
                    cell.value = ""  # 非System/User Requirementも空欄
                    cell.font = Font(name=self.DEFAULT_FONT_NAME)
            else:
                cell = ws[f"W{row_num}"]
                cell.value = ""  # Descriptionがない場合も空欄
                cell.font = Font(name=self.DEFAULT_FONT_NAME)
                    
            row_num += 1
            
        logger.info("要件シート作成完了")
        
        # 列幅調整
        logger.info("列幅調整中...")
        self._adjust_column_widths(ws)
        logger.info("列幅調整完了")
        
        # 新規要件追加用の説明行を追加
        self._add_new_item_instructions(ws, row_num)
        
    def _add_new_item_instructions(self, ws, start_row: int) -> None:
        """新規要件追加用の説明を追加"""
        # 空行を挿入
        start_row += 2
        
        # 説明テキスト
        instructions = [
            "【新規要件の追加方法】",
            "1. 上記の最終行の下に新しい行を追加",
            "2. A列（JAMA_ID）は空欄のまま",
            "3. D～N列（階層1～11）に配置したい階層名を入力",
            "4. 新規System Requirementの場合：N列に要件名を入力し、X列に「#S1」～「#S200」を入力",
            "5. 新規User Requirementの場合：N列に要件名を入力し、X列に「#U1」～「#U200」を入力",
            "6. その他の必要な情報を入力",
            "7. updateコマンドを実行"
        ]
        
        for i, instruction in enumerate(instructions):
            cell = ws[f"A{start_row + i}"]
            cell.value = instruction
            cell.font = Font(name=self.DEFAULT_FONT_NAME, color="666666", italic=True)
            
        # 新規要件用の説明
        cell = ws[f"X{start_row + 4}"]
        cell.value = "新規System Req: #S1～#S200"
        cell.font = Font(name=self.DEFAULT_FONT_NAME, color="FF0000", italic=True)
        
        cell = ws[f"X{start_row + 5}"]
        cell.value = "新規User Req: #U1～#U200"
        cell.font = Font(name=self.DEFAULT_FONT_NAME, color="FF0000", italic=True)
        
    def _create_user_description_sheet(self, items: List[Dict]) -> None:
        """User Description編集シートを作成"""
        ws = self.user_description_sheet
        
        # スタイル定義
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # User Requirementアイテムをフィルタリング（ItemType ID: 266）
        logger.info("STD User Requirement (ID:266)アイテムをフィルタリング中...")
        user_requirement_items = []
        for idx, item in enumerate(items):
            # item_type_idが266かチェック
            if item.get("item_type_id") == 266:
                user_requirement_items.append((item, idx))
        
        logger.info(f"STD User Requirement (ID:266)アイテム数: {len(user_requirement_items)}件")
        
        if not user_requirement_items:
            logger.info("STD User Requirement (ID:266)アイテムが見つかりませんでした")
            cell = ws["A1"]
            cell.value = "STD User Requirement (ID:266)アイテムが見つかりませんでした"
            cell.font = Font(name=self.DEFAULT_FONT_NAME)
        else:
            # 既存のUser Requirementアイテムにテンプレートを作成
            current_row = 10
            total_user_requirement = len(user_requirement_items)
            
            logger.info(f"User Descriptionテンプレート作成開始: {total_user_requirement}件")
            
            # User Requirementアイテムのインデックスを保存（リンク作成用）
            self.user_requirement_template_map = {}  # row_index -> template_start_row
            
            for idx, (item, original_idx) in enumerate(user_requirement_items, 1):
                # 進捗表示
                if idx % 100 == 0 or idx == total_user_requirement:
                    logger.info(f"User Descriptionテンプレート作成進捗: {idx}/{total_user_requirement} ({idx/total_user_requirement*100:.1f}%)")
                    
                # マッピング情報を保存
                self.user_requirement_template_map[original_idx] = current_row
                
                # ヘッダー行（【JAMA_ID】要件名の形式で表示）
                jama_id = item.get('jama_id', '新規')
                name = item.get("name", "")
                cell = ws[f"A{current_row - 2}"]
                cell.value = f"【{jama_id}】{name}"
                cell.font = Font(name=self.DEFAULT_FONT_NAME, bold=True)
                # 列の結合を動的に計算（最後の列を残す）
                merge_end_col = get_column_letter(self.user_desc_total_cols - 1)
                ws.merge_cells(f"A{current_row - 2}:{merge_end_col}{current_row - 2}")
                
                # 要件名を表示（リンク先の1つ上）
                cell = ws[f"A{current_row - 1}"]
                cell.value = name
                cell.font = Font(name=self.DEFAULT_FONT_NAME, bold=True, size=12)
                ws.merge_cells(f"A{current_row - 1}:{merge_end_col}{current_row - 1}")
                
                # 一覧に戻るリンク（最後の列）
                link_col = get_column_letter(self.user_desc_total_cols)
                cell = ws[f"{link_col}{current_row - 1}"]
                cell.value = "一覧に戻る"
                cell.hyperlink = f"#Requirement_of_Driver!A{original_idx + 2}"
                cell.font = Font(name=self.DEFAULT_FONT_NAME, color="0000FF", underline="single")
                
                # 既存のDescriptionをパース
                existing_data = None
                if item.get("description"):
                    existing_data = self._parse_existing_user_description_table(item.get("description"))
                    if existing_data:
                        logger.info(f"既存のDescriptionテーブルを検出: JAMA_ID={jama_id}")
                
                # 3行テーブルのテンプレート作成（既存データがあれば渡す）
                self._create_user_description_template(ws, current_row, existing_data)
                
                current_row += 12  # 次のテンプレートまでの間隔（3行テーブルなので間隔を調整）
                
            logger.info(f"User Descriptionテンプレート作成完了: {total_user_requirement}件")
        
        # 新規User Requirement用のテンプレートを追加
        logger.info(f"新規User Requirement用テンプレート作成開始: {self.NEW_USER_REQUIREMENT_TEMPLATE_COUNT}件")
        self._create_new_user_requirement_templates(ws, current_row if 'current_row' in locals() else 10)
        logger.info("新規User Requirement用テンプレート作成完了")
        
        # 列幅設定（固定幅）
        logger.info("User Description編集シートの列幅設定中...")
        self._set_fixed_column_widths(ws, self.user_desc_total_cols)
        logger.info("列幅設定完了")
        
    def _create_system_description_sheet(self, items: List[Dict]) -> None:
        """System Description編集シートを作成"""
        ws = self.system_description_sheet
        
        # スタイル定義
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # System Requirementアイテムをフィルタリング（ItemType ID: 301）
        logger.info("STD Oneteam System Requirement (ID:301)アイテムをフィルタリング中...")
        system_requirement_items = []
        for idx, item in enumerate(items):
            # item_type_idが301かチェック
            if item.get("item_type_id") == 301:
                system_requirement_items.append((item, idx))
        
        logger.info(f"STD Oneteam System Requirement (ID:301)アイテム数: {len(system_requirement_items)}件")
        
        if not system_requirement_items:
            logger.info("STD Oneteam System Requirement (ID:301)アイテムが見つかりませんでした")
            cell = ws["A1"]
            cell.value = "STD Oneteam System Requirement (ID:301)アイテムが見つかりませんでした"
            cell.font = Font(name=self.DEFAULT_FONT_NAME)
        else:
            # 既存のSystem Requirementアイテムにテンプレートを作成
            current_row = 10
            total_system_requirement = len(system_requirement_items)
            
            logger.info(f"System Descriptionテンプレート作成開始: {total_system_requirement}件")
            
            # System Requirementアイテムのインデックスを保存（リンク作成用）
            self.system_requirement_template_map = {}  # row_index -> template_start_row
            
            for idx, (item, original_idx) in enumerate(system_requirement_items, 1):
                # 進捗表示
                if idx % 100 == 0 or idx == total_system_requirement:
                    logger.info(f"System Descriptionテンプレート作成進捗: {idx}/{total_system_requirement} ({idx/total_system_requirement*100:.1f}%)")
                    
                # マッピング情報を保存
                self.system_requirement_template_map[original_idx] = current_row
                
                # ヘッダー行（【JAMA_ID】要件名の形式で表示）
                jama_id = item.get('jama_id', '新規')
                name = item.get("name", "")
                cell = ws[f"A{current_row - 2}"]
                cell.value = f"【{jama_id}】{name}"
                cell.font = Font(name=self.DEFAULT_FONT_NAME, bold=True)
                # 列の結合を動的に計算
                merge_end_col = get_column_letter(min(10, self.system_desc_total_cols))
                ws.merge_cells(f"A{current_row - 2}:{merge_end_col}{current_row - 2}")
                
                # 要件名を表示（リンク先の1つ上）
                cell = ws[f"A{current_row - 1}"]
                cell.value = name
                cell.font = Font(name=self.DEFAULT_FONT_NAME, bold=True, size=12)
                ws.merge_cells(f"A{current_row - 1}:{merge_end_col}{current_row - 1}")
                
                # 一覧に戻るリンク
                link_col = get_column_letter(min(11, self.system_desc_total_cols))
                cell = ws[f"{link_col}{current_row - 1}"]
                cell.value = "一覧に戻る"
                cell.hyperlink = f"#Requirement_of_Driver!A{original_idx + 2}"
                cell.font = Font(name=self.DEFAULT_FONT_NAME, color="0000FF", underline="single")
                
                # 既存のDescriptionをパース
                existing_data = None
                if item.get("description"):
                    existing_data = self._parse_existing_description_table(item.get("description"))
                    if existing_data:
                        logger.info(f"既存のDescriptionテーブルを検出: JAMA_ID={jama_id}")
                
                # 5行テーブルのテンプレート作成（既存データがあれば渡す）
                self._create_description_template(ws, current_row, existing_data)
                
                current_row += 16  # 次のテンプレートまでの間隔
                
            logger.info(f"System Descriptionテンプレート作成完了: {total_system_requirement}件")
        
        # 新規System Requirement用のテンプレートを追加
        logger.info(f"新規System Requirement用テンプレート作成開始: {self.NEW_SYSTEM_REQUIREMENT_TEMPLATE_COUNT}件")
        self._create_new_system_requirement_templates(ws, current_row if 'current_row' in locals() else 10)
        logger.info("新規System Requirement用テンプレート作成完了")
        
        # 列幅設定（固定幅）
        logger.info("System Description編集シートの列幅設定中...")
        self._set_fixed_column_widths(ws, self.system_desc_total_cols)
        logger.info("列幅設定完了")
        
    def _create_new_user_requirement_templates(self, ws, start_row: int) -> None:
        """新規User Requirement用の空テンプレートを作成"""
        # セクションタイトル
        cell = ws[f"A{start_row}"]
        cell.value = "=== 新規STD User Requirement (ID:266)追加用テンプレート ==="
        cell.font = Font(name=self.DEFAULT_FONT_NAME, bold=True, size=14, color="FF0000")
        merge_end_col = get_column_letter(min(20, self.user_desc_total_cols))
        ws.merge_cells(f"A{start_row}:{merge_end_col}{start_row}")
        
        start_row += 3
        
        # 新規テンプレートを作成
        for i in range(1, self.NEW_USER_REQUIREMENT_TEMPLATE_COUNT + 1):
            # 進捗表示（50個ごと）
            if i % 50 == 0:
                logger.info(f"新規User Requirementテンプレート作成進捗: {i}/{self.NEW_USER_REQUIREMENT_TEMPLATE_COUNT}")
                
            # ヘッダー行
            cell = ws[f"A{start_row}"]
            cell.value = f"【新規User Requirement #U{i}】ここに要件名を入力"
            cell.font = Font(name=self.DEFAULT_FONT_NAME, bold=True, color="FF0000")
            merge_end_col = get_column_letter(self.user_desc_total_cols - 1)
            ws.merge_cells(f"A{start_row}:{merge_end_col}{start_row}")
            
            # 使用方法の説明
            cell = ws[f"A{start_row + 1}"]
            cell.value = f"Requirement_of_DriverシートのX列に「#U{i}」と入力してこのテンプレートを使用"
            cell.font = Font(name=self.DEFAULT_FONT_NAME, italic=True, color="666666", size=9)
            ws.merge_cells(f"A{start_row + 1}:{merge_end_col}{start_row + 1}")
            
            # 空のテンプレート作成
            self._create_user_description_template(ws, start_row + 3, None)
            
            start_row += 12  # 次のテンプレートまでの間隔（3行テーブルなので間隔を調整）
                    
    def _create_new_system_requirement_templates(self, ws, start_row: int) -> None:
        """新規System Requirement用の空テンプレートを作成"""
        # セクションタイトル
        cell = ws[f"A{start_row}"]
        cell.value = "=== 新規STD Oneteam System Requirement (ID:301)追加用テンプレート ==="
        cell.font = Font(name=self.DEFAULT_FONT_NAME, bold=True, size=14, color="FF0000")
        merge_end_col = get_column_letter(min(20, self.system_desc_total_cols))
        ws.merge_cells(f"A{start_row}:{merge_end_col}{start_row}")
        
        start_row += 3
        
        # 新規テンプレートを作成
        for i in range(1, self.NEW_SYSTEM_REQUIREMENT_TEMPLATE_COUNT + 1):
            # 進捗表示（50個ごと）
            if i % 50 == 0:
                logger.info(f"新規System Requirementテンプレート作成進捗: {i}/{self.NEW_SYSTEM_REQUIREMENT_TEMPLATE_COUNT}")
                
            # ヘッダー行
            cell = ws[f"A{start_row}"]
            cell.value = f"【新規System Requirement #S{i}】ここに要件名を入力"
            cell.font = Font(name=self.DEFAULT_FONT_NAME, bold=True, color="FF0000")
            merge_end_col = get_column_letter(min(10, self.system_desc_total_cols))
            ws.merge_cells(f"A{start_row}:{merge_end_col}{start_row}")
            
            # 使用方法の説明
            cell = ws[f"A{start_row + 1}"]
            cell.value = f"Requirement_of_DriverシートのX列に「#S{i}」と入力してこのテンプレートを使用"
            cell.font = Font(name=self.DEFAULT_FONT_NAME, italic=True, color="666666", size=9)
            ws.merge_cells(f"A{start_row + 1}:{merge_end_col}{start_row + 1}")
            
            # 空のテンプレート作成
            self._create_description_template(ws, start_row + 3, None)
            
            start_row += 16  # 次のテンプレートまでの間隔
            
    def _create_user_description_template(self, ws, start_row: int, existing_data: Optional[List[List[str]]] = None) -> None:
        """
        3行形式のUser Descriptionテンプレートを作成（拡張性対応版）
        
        Args:
            ws: ワークシート
            start_row: 開始行
            existing_data: 既存のテーブルデータ（3行のリスト）
        """
        # 列の位置を動的に計算
        pos = self._get_user_column_positions()
        
        # スタイル
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # 1行目: I/O Type
        cell = ws.cell(row=start_row, column=1, value="I/O Type")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        cell = ws.cell(row=start_row, column=2, value="IN")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        cell = ws.cell(row=start_row, column=3, value="OUT")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        # OUTを最後まで結合（動的に計算）
        end_col = get_column_letter(self.user_desc_total_cols)
        ws.merge_cells(f"C{start_row}:{end_col}{start_row}")
        
        # 2行目: 項目名
        cell = ws.cell(row=start_row + 1, column=1, value="")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        # (a) Trigger action
        cell = ws.cell(row=start_row + 1, column=pos['a_start'], value="(a)Trigger action")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        # (c) HMI（動的に結合）
        cell = ws.cell(row=start_row + 1, column=pos['c_start'], value="(c)HMI")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        if pos['c_end'] > pos['c_start']:
            ws.merge_cells(f"{get_column_letter(pos['c_start'])}{start_row + 1}:{get_column_letter(pos['c_end'])}{start_row + 1}")
            
        # 3行目: 要件
        cell = ws.cell(row=start_row + 2, column=1, value="要件")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        # 入力セルを配置（動的に範囲を計算）
        for col in range(2, self.user_desc_total_cols + 1):
            cell = ws.cell(row=start_row + 2, column=col)
            
            # 既存データがある場合は値を設定
            if existing_data and len(existing_data) > 2 and col <= len(existing_data[2]):
                cell.value = existing_data[2][col - 1]
            else:
                cell.value = ""
                
            cell.border = border
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(name=self.DEFAULT_FONT_NAME)
            
    def _create_description_template(self, ws, start_row: int, existing_data: Optional[List[List[str]]] = None) -> None:
        """
        5行形式のDescriptionテンプレートを作成（拡張性対応版）
        
        Args:
            ws: ワークシート
            start_row: 開始行
            existing_data: 既存のテーブルデータ（5行のリスト）
        """
        # 列の位置を動的に計算
        pos = self._get_system_column_positions()
        
        # スタイル
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # 1行目: I/O Type
        cell = ws.cell(row=start_row, column=1, value="I/O Type")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        cell = ws.cell(row=start_row, column=2, value="IN")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        cell = ws.cell(row=start_row, column=3, value="OUT")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        # OUTを最後まで結合（動的に計算）
        end_col = get_column_letter(self.system_desc_total_cols)
        ws.merge_cells(f"C{start_row}:{end_col}{start_row}")
        
        # 2行目: 項目名
        cell = ws.cell(row=start_row + 1, column=1, value="")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        # (a) Trigger action
        cell = ws.cell(row=start_row + 1, column=pos['a_start'], value="(a)Trigger action")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        
        # (b) Behavior of ego-vehicle（動的に結合）
        cell = ws.cell(row=start_row + 1, column=pos['b_start'], value="(b)Behavior of ego-vehicle")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        if pos['b_end'] > pos['b_start']:
            ws.merge_cells(f"{get_column_letter(pos['b_start'])}{start_row + 1}:{get_column_letter(pos['b_end'])}{start_row + 1}")
            
        # (c) HMI（動的に結合）
        cell = ws.cell(row=start_row + 1, column=pos['c_start'], value="(c)HMI")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        if pos['c_end'] > pos['c_start']:
            ws.merge_cells(f"{get_column_letter(pos['c_start'])}{start_row + 1}:{get_column_letter(pos['c_end'])}{start_row + 1}")
            
        # (d) Other（動的に結合）
        cell = ws.cell(row=start_row + 1, column=pos['d_start'], value="(d)Other")
        cell.border = border
        cell.font = Font(name=self.DEFAULT_FONT_NAME)
        if pos['d_end'] > pos['d_start']:
            ws.merge_cells(f"{get_column_letter(pos['d_start'])}{start_row + 1}:{get_column_letter(pos['d_end'])}{start_row + 1}")
            
        # 3-5行目: Data Name, Data Label, Data
        data_rows = ["Data Name", "Data Label", "Data"]
        for i, row_name in enumerate(data_rows, 2):
            cell = ws.cell(row=start_row + i, column=1, value=row_name)
            cell.border = border
            cell.font = Font(name=self.DEFAULT_FONT_NAME)
            
            # 入力セルを配置（動的に範囲を計算）
            for col in range(2, self.system_desc_total_cols + 1):
                cell = ws.cell(row=start_row + i, column=col)
                
                # 既存データがある場合は値を設定
                if existing_data and i >= 2 and col <= len(existing_data[i]):
                    # existing_dataのインデックスは0ベース、iは2から始まる
                    cell.value = existing_data[i][col - 1]
                else:
                    cell.value = ""
                    
                cell.border = border
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                cell.font = Font(name=self.DEFAULT_FONT_NAME)
                
    def _parse_hierarchy_fast(self, item: Dict) -> List[str]:
        """
        アイテムの階層構造を高速に解析（インデックス使用）
        
        Args:
            item: 対象アイテム
            
        Returns:
            階層名のリスト
        """
        hierarchy = []
        sequence = item.get("sequence", "")
        
        if not sequence:
            return [item.get("name", "")]
            
        # sequence の各レベルでアイテムを探す（インデックス使用で高速化）
        parts = sequence.split(".")
        for i in range(1, len(parts) + 1):
            current_seq = ".".join(parts[:i])
            # インデックスから直接取得
            if current_seq in self.sequence_index:
                hierarchy.append(self.sequence_index[current_seq].get("name", ""))
                    
        return hierarchy
        
    def _parse_hierarchy(self, item: Dict, all_items: List[Dict]) -> List[str]:
        """
        アイテムの階層構造を解析
        
        Args:
            item: 対象アイテム
            all_items: 全アイテムリスト
            
        Returns:
            階層名のリスト
        """
        hierarchy = []
        sequence = item.get("sequence", "")
        
        if not sequence:
            return [item.get("name", "")]
            
        # sequence の各レベルでアイテムを探す
        parts = sequence.split(".")
        for i in range(1, len(parts) + 1):
            current_seq = ".".join(parts[:i])
            # 該当するアイテムを探す
            for other_item in all_items:
                if other_item.get("sequence") == current_seq:
                    hierarchy.append(other_item.get("name", ""))
                    break
                    
        return hierarchy
        
    def _extract_table_preview(self, html_description: str) -> str:
        """
        HTMLテーブルから簡易プレビューを作成
        
        Args:
            html_description: HTML形式のDescription
            
        Returns:
            プレビュー文字列
        """
        if not html_description:
            return ""
            
        # HTMLパーサーでテーブルを抽出
        parser = HTMLTableParser()
        parser.feed(html_description)
        
        if parser.tables:
            # 最初のテーブルの最初の3行を表示
            table = parser.tables[0]
            preview_rows = table[:3]
            preview = []
            for row in preview_rows:
                preview.append(" | ".join(row[:4]))  # 最初の4列のみ
            return "\n".join(preview)
            
        # テーブルがない場合はテキストを抽出
        text = re.sub('<[^<]+?>', '', html_description)
        return text[:100] + "..." if len(text) > 100 else text
        
    def _parse_existing_description_table(self, html_description: str) -> Optional[List[List[str]]]:
        """
        既存のHTML Descriptionテーブルをパースして値を取得
        
        Args:
            html_description: HTML形式のDescription
            
        Returns:
            5行×N列のテーブルデータ、または None（形式が異なる場合）
        """
        if not html_description or '<table' not in html_description:
            return None
            
        # HTMLパーサーでテーブルを抽出
        parser = HTMLTableParser()
        parser.feed(html_description)
        
        if not parser.tables:
            return None
            
        table = parser.tables[0]
        
        # テーブル構造の検証
        if len(table) != 5:
            return None
            
        # 固定ヘッダーの検証
        # 1行目
        if len(table[0]) < 3:
            return None
        if table[0][0] != "I/O Type" or table[0][1] != "IN" or table[0][2] != "OUT":
            return None
            
        # 2行目
        if len(table[1]) < 5:
            return None
        if table[1][0] != "" or table[1][1] != "(a)Trigger action":
            return None
        if "(b)Behavior of ego-vehicle" not in table[1][2]:
            return None
        # 2行目の後半のヘッダーは、HTMLパーサーの結果によって位置が異なる可能性があるため
        # より柔軟な検証にする
        row2_text = " ".join(table[1])
        if "(c)HMI" not in row2_text or "(d)Other" not in row2_text:
            return None
            
        # 3-5行目
        if table[2][0] != "Data Name" or table[3][0] != "Data Label" or table[4][0] != "Data":
            return None
            
        # 構造が正しいことを確認したら、値を返す
        return table
        
    def _parse_existing_user_description_table(self, html_description: str) -> Optional[List[List[str]]]:
        """
        既存のHTML User Descriptionテーブルをパースして値を取得
        
        Args:
            html_description: HTML形式のDescription
            
        Returns:
            3行×N列のテーブルデータ、または None（形式が異なる場合）
        """
        if not html_description or '<table' not in html_description:
            return None
            
        # HTMLパーサーでテーブルを抽出
        parser = HTMLTableParser()
        parser.feed(html_description)
        
        if not parser.tables:
            return None
            
        table = parser.tables[0]
        
        # User Description用の3行フォーマットを試みる
        # まず5行フォーマットかチェック
        if len(table) == 5:
            # 5行フォーマットの場合、3行フォーマットに変換を試みる
            # 3-5行目のデータを統合
            converted_data = []
            converted_data.append(table[0])  # I/O Type行
            converted_data.append(table[1])  # 項目名行
            
            # データ行を統合（Data Name, Data Label, Dataを結合）
            data_row = ["要件"]
            # 各列のデータを結合
            for col_idx in range(1, len(table[2])):
                combined = []
                if col_idx < len(table[2]) and table[2][col_idx]:
                    combined.append(table[2][col_idx])
                if col_idx < len(table[3]) and table[3][col_idx]:
                    combined.append(table[3][col_idx])
                if col_idx < len(table[4]) and table[4][col_idx]:
                    combined.append(table[4][col_idx])
                data_row.append(" / ".join(combined))
            
            converted_data.append(data_row)
            return converted_data
            
        # 3行フォーマットの場合
        if len(table) == 3:
            # 固定ヘッダーの検証
            # 1行目
            if len(table[0]) < 3:
                return None
            if table[0][0] != "I/O Type" or table[0][1] != "IN" or table[0][2] != "OUT":
                return None
                
            # 構造が正しいことを確認したら、値を返す
            return table
            
        # その他のフォーマットは対応しない
        return None
    
    def _adjust_column_widths(self, ws, max_columns: Optional[int] = None) -> None:
        """
        列幅を自動調整
        
        Args:
            ws: ワークシート
            max_columns: 調整する最大列数（Noneの場合はCOLUMNSの数）
        """
        if max_columns is None:
            total_columns = len(self.COLUMNS)
        else:
            total_columns = max_columns
            
        logger.info(f"列幅自動調整開始: {total_columns}列")
        
        for idx, column in enumerate(ws.columns, 1):
            if idx > total_columns:
                break
                
            # 進捗表示
            if idx % 5 == 0 or idx == total_columns:
                logger.info(f"列幅調整進捗: {idx}/{total_columns}")
                
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # 設定値に基づいてチェックする行数を制限（高速化のため）
            check_rows = min(self.column_width_check_rows, len(column))
            for cell in list(column)[:check_rows]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)  # 最大幅を50に制限
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _set_fixed_column_widths(self, ws, total_columns: int) -> None:
        """
        固定列幅を設定
        
        Args:
            ws: ワークシート
            total_columns: 設定する列数
        """
        # A列
        ws.column_dimensions['A'].width = self.FIXED_COLUMN_WIDTH_A
        
        # B列以降
        for col_idx in range(2, total_columns + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = self.FIXED_COLUMN_WIDTH_OTHER
            
    def read_requirement_excel(self, input_file: str) -> List[Dict]:
        """
        Excelファイルから要件データを読み込み（改良版）

        Args:
            input_file: 入力ファイル名

        Returns:
            要件データリスト（処理対象のみ）
        """
        logger.info(f"Excelファイル読み込み開始: {input_file}")
    
        try:
            # ファイルサイズをチェック
            file_size = os.path.getsize(input_file) / 1024  # KB単位
            logger.info(f"ファイルサイズ: {file_size:.1f}KB")
        
            # 読み込み開始メッセージ
            print(f"\nExcelファイルを開いています... (ファイルサイズ: {file_size:.1f}KB)")
            print("5MB程度のファイルで約2分かかる場合があります。しばらくお待ちください...")
        
            # ファイルを読み込み
            wb = openpyxl.load_workbook(input_file)
            print("Excelファイルを開きました。データを処理しています...")
        
            requirement_sheet = wb["Requirement_of_Driver"]
            
            # Description編集シートの存在確認
            system_description_sheet = wb["System_Description_edit"] if "System_Description_edit" in wb.sheetnames else None
            user_description_sheet = wb["User_Description_edit"] if "User_Description_edit" in wb.sheetnames else None
        
            requirements = []
            total_rows = requirement_sheet.max_row - 1  # ヘッダー行を除く
        
            logger.info(f"総行数: {total_rows}")
            
            # ステップ1: すべての既存要件を読み込み（バリデーション用）
            logger.info("既存要件の読み込み開始...")
            self.all_existing_items = []  # インスタンス変数として保持
            
            for row_num in range(2, requirement_sheet.max_row + 1):
                jama_id = requirement_sheet[f"A{row_num}"].value
                if jama_id:  # 既存要件（JAMA_IDがある）
                    # すべての情報を読み込み
                    item = {
                        "jama_id": jama_id,
                        "sequence": requirement_sheet[f"C{row_num}"].value or "",  # Noneを空文字列に変換
                        "row_num": row_num,
                        "name": requirement_sheet[f"N{row_num}"].value or requirement_sheet[f"M{row_num}"].value
                    }
                    
                    # 階層情報も保持
                    hierarchy = []
                    for col in range(4, 15):  # D～N列
                        value = requirement_sheet.cell(row=row_num, column=col).value
                        if value:
                            hierarchy.append(str(value).strip())
                        else:
                            break
                    item["hierarchy"] = hierarchy
                    
                    self.all_existing_items.append(item)
                    
            logger.info(f"既存要件読み込み完了: {len(self.all_existing_items)}件")
        
            # ステップ2: 処理対象の要件を読み込み
            logger.info("処理対象要件の読み込み開始...")
            
            for row_idx, row_num in enumerate(range(2, requirement_sheet.max_row + 1), 1):
                # 進捗表示（100行ごと、最初と最後）
                if row_idx == 1 or row_idx % 100 == 0 or row_idx == total_rows:
                    logger.info(f"処理対象要件読み込み進捗: {row_idx}/{total_rows} ({row_idx/total_rows*100:.1f}%)")
            
                # 基本情報を取得
                jama_id = requirement_sheet[f"A{row_num}"].value
                update_flag = requirement_sheet[f"W{row_num}"].value
                memo_comment = requirement_sheet[f"B{row_num}"].value
                sequence = requirement_sheet[f"C{row_num}"].value
                
                # 空欄行のチェック（新規要件と区別）
                if not jama_id:  # JAMA_IDが空欄
                    # 主要なデータ列をチェック
                    has_data = False
                    # 階層情報（D～N列）をチェック
                    for col in range(4, 15):
                        if requirement_sheet.cell(row=row_num, column=col).value:
                            has_data = True
                            break
                    
                    # その他の重要フィールドもチェック
                    if not has_data:
                        for col in ['P', 'Q', 'R', 'S', 'T', 'U']:
                            if requirement_sheet[f"{col}{row_num}"].value:
                                has_data = True
                                break
                                
                    if not has_data:
                        # 完全な空欄行は静かにスキップ
                        continue
            
                # 削除操作の判定（B列に「削除」と記載）- 最優先
                if str(memo_comment).strip() == "削除" and jama_id:
                    operation = "削除"
                # 処理対象の判定
                elif not jama_id:
                    # JAMA_IDが空欄だが、データがある場合は新規作成
                    operation = "新規"
                elif update_flag == "する":
                    operation = "更新"
                else:
                    # JAMA_IDはあるが更新フラグが「する」でない場合はスキップ
                    continue
                
                # 基本情報を読み込み
                item = {
                    "operation": operation,
                    "sequence": sequence,
                    "name": requirement_sheet[f"N{row_num}"].value or requirement_sheet[f"M{row_num}"].value,  # 階層11または10
                }
            
                # JAMA IDがある場合のみ設定
                if jama_id:
                    item["jama_id"] = jama_id
                    
                # 新規作成の場合、階層から親要件とsequenceを計算
                if operation == "新規":
                    hierarchy_info = self._calculate_hierarchy_info_improved(requirement_sheet, row_num)
                    if hierarchy_info:
                        item.update(hierarchy_info)
                    else:
                        # エラーメッセージは_calculate_hierarchy_info_improved内で出力済み
                        continue
            
                # その他のフィールドは、値がある場合のみ設定
                field_mapping = {
                    "P": "assignee",
                    "Q": "status",
                    "R": "tags",
                    "S": "reason",
                    "T": "preconditions",
                    "U": "target_system"
                }
            
                for col, field_name in field_mapping.items():
                    value = requirement_sheet[f"{col}{row_num}"].value
                    if value is not None and value != "":
                        item[field_name] = value
            
                # Description更新チェック
                if (update_flag == "する" or operation == "新規"):
                    # X列のDescription参照を確認
                    desc_ref = requirement_sheet[f"X{row_num}"].value
                    
                    if desc_ref:
                        # プレフィックスで判定
                        desc_ref_str = str(desc_ref)
                        
                        if desc_ref_str.startswith("#S"):
                            # System Requirement用テンプレートから読み込み
                            if system_description_sheet:
                                new_description = self._read_new_requirement_description(
                                    system_description_sheet,
                                    desc_ref_str,
                                    "System"
                                )
                                if new_description:
                                    item["description"] = new_description
                                    
                        elif desc_ref_str.startswith("#U"):
                            # User Requirement用テンプレートから読み込み
                            if user_description_sheet:
                                new_description = self._read_new_requirement_description(
                                    user_description_sheet,
                                    desc_ref_str,
                                    "User"
                                )
                                if new_description:
                                    item["description"] = new_description
                                    
                        else:
                            # プレフィックスがない場合（既存要件の編集）
                            # JAMA IDからアイテムタイプを判定して適切なシートを選択
                            if jama_id and operation == "更新":
                                # ここでアイテムタイプを判定する方法が必要
                                # 仮実装：「編集画面へ」リンクのターゲットシートで判定
                                if system_description_sheet:
                                    new_description = self._read_description_from_sheet(
                                        system_description_sheet,
                                        jama_id,
                                        "System"
                                    )
                                    if new_description:
                                        item["description"] = new_description
                                    elif user_description_sheet:
                                        # System Descriptionに見つからない場合はUser Descriptionを確認
                                        new_description = self._read_description_from_sheet(
                                            user_description_sheet,
                                            jama_id,
                                            "User"
                                        )
                                        if new_description:
                                            item["description"] = new_description
                        
                requirements.append(item)
            
            logger.info(f"読み込み完了: {len(requirements)}件")
            logger.info(f"  新規作成: {len([r for r in requirements if r['operation'] == '新規'])}件")
            logger.info(f"  更新: {len([r for r in requirements if r['operation'] == '更新'])}件")
            logger.info(f"  削除: {len([r for r in requirements if r['operation'] == '削除'])}件")
        
            return requirements
        
        except Exception as e:
            logger.error(f"Excelファイル読み込みエラー: {str(e)}")
            raise
            
    def _calculate_hierarchy_info_improved(self, ws, row_num: int) -> Optional[Dict]:
        """
        新規要件の階層情報から親要件とsequenceを計算（改良版）
        
        Args:
            ws: ワークシート
            row_num: 行番号
            
        Returns:
            階層情報の辞書、またはNone（エラーの場合）
        """
        # C列（Sequence）が指定されている場合は優先使用
        manual_sequence = ws[f"C{row_num}"].value
        if manual_sequence:
            logger.info(f"行{row_num}: C列にSequence指定あり: {manual_sequence}")
            
            # 親Sequenceを計算
            seq_parts = str(manual_sequence).split(".")
            if len(seq_parts) == 1:
                # ルート直下
                return {
                    "name": ws[f"N{row_num}"].value or ws[f"M{row_num}"].value,
                    "parent_id": None,
                    "calculated_sequence": manual_sequence
                }
            else:
                # 親Sequenceを構築
                parent_sequence = ".".join(seq_parts[:-1])
                # 親を検索
                parent_item = None
                for item in self.all_existing_items:
                    if item.get("sequence") == parent_sequence:
                        parent_item = item
                        break
                        
                if parent_item:
                    return {
                        "name": ws[f"N{row_num}"].value or ws[f"M{row_num}"].value,
                        "parent_id": parent_item["jama_id"],
                        "calculated_sequence": manual_sequence
                    }
                else:
                    logger.error(f"行{row_num}: 指定されたSequence {manual_sequence} の親 {parent_sequence} が見つかりません")
                    return None
        
        # C列が空欄の場合、階層情報から推定
        # D～N列（階層1～11）から階層情報を取得
        hierarchy = []
        for col in range(4, 15):  # D=4, N=14
            value = ws.cell(row=row_num, column=col).value
            if value:
                hierarchy.append(str(value).strip())
            else:
                break
                
        if not hierarchy:
            logger.error(f"行{row_num}: 階層情報が設定されていません")
            return None
            
        # 最も深い階層が新規要件の名前
        name = hierarchy[-1]
        
        # 親階層を探す（最後の要素を除く）
        if len(hierarchy) == 1:
            # ルート直下の場合
            return {
                "name": name,
                "parent_id": None,
                "calculated_sequence": self._get_next_sequence("", self.all_existing_items)
            }
        else:
            # 親階層の候補を検索
            parent_hierarchy = hierarchy[:-1]
            parent_candidates = self._find_parent_candidates(parent_hierarchy)
            
            if len(parent_candidates) == 0:
                # 親が見つからない場合、類似候補を提案
                similar_candidates = self._find_similar_parents(parent_hierarchy)
                error_msg = f"行{row_num}: 親要件が見つかりません\n"
                error_msg += f"  探している親階層: {' > '.join(parent_hierarchy)}\n"
                
                if similar_candidates:
                    error_msg += "  類似する候補:\n"
                    for candidate in similar_candidates[:3]:  # 最大3件表示
                        error_msg += f"    - [{candidate['jama_id']}] {' > '.join(candidate['hierarchy'])}\n"
                    error_msg += "  C列にSequenceを入力して明示的に指定してください"
                else:
                    error_msg += "  類似する候補も見つかりませんでした"
                    
                logger.error(error_msg)
                print(f"\n❌ {error_msg}")
                return None
                
            elif len(parent_candidates) == 1:
                # 親が1つだけ見つかった場合
                parent_item = parent_candidates[0]
                parent_sequence = parent_item["sequence"]
                return {
                    "name": name,
                    "parent_id": parent_item["jama_id"],
                    "calculated_sequence": self._get_next_sequence(parent_sequence, self.all_existing_items)
                }
                
            else:
                # 複数の候補が見つかった場合
                error_msg = f"行{row_num}: 親要件が複数見つかりました\n"
                error_msg += f"  探している親階層: {' > '.join(parent_hierarchy)}\n"
                error_msg += "  候補:\n"
                for idx, candidate in enumerate(parent_candidates):
                    error_msg += f"    {idx+1}. [{candidate['jama_id']}] "
                    error_msg += f"Sequence: {candidate['sequence']}, "
                    error_msg += f"階層: {' > '.join(candidate['hierarchy'])}\n"
                error_msg += "  C列にSequenceを入力して明示的に指定してください"
                
                logger.error(error_msg)
                print(f"\n❌ {error_msg}")
                return None
    
    def _find_parent_candidates(self, parent_hierarchy: List[str]) -> List[Dict]:
        """
        階層名から親要件の候補を検索
        
        Args:
            parent_hierarchy: 親階層名のリスト
            
        Returns:
            候補のリスト
        """
        candidates = []
        
        for item in self.all_existing_items:
            if item["hierarchy"] == parent_hierarchy:
                candidates.append(item)
                
        return candidates
        
    def _find_similar_parents(self, parent_hierarchy: List[str]) -> List[Dict]:
        """
        類似する親要件を検索
        
        Args:
            parent_hierarchy: 親階層名のリスト
            
        Returns:
            類似候補のリスト
        """
        similar = []
        target_depth = len(parent_hierarchy)
        
        for item in self.all_existing_items:
            item_hierarchy = item["hierarchy"]
            if len(item_hierarchy) != target_depth:
                continue
                
            # 最後の要素が一致する候補を探す
            if item_hierarchy[-1] == parent_hierarchy[-1]:
                similar.append(item)
                
        # 類似度でソート（後ろから一致する要素数が多い順）
        similar.sort(key=lambda x: self._calculate_similarity(x["hierarchy"], parent_hierarchy), reverse=True)
        
        return similar
        
    def _calculate_similarity(self, hierarchy1: List[str], hierarchy2: List[str]) -> int:
        """
        2つの階層の類似度を計算
        
        Args:
            hierarchy1: 階層1
            hierarchy2: 階層2
            
        Returns:
            類似度スコア
        """
        score = 0
        min_len = min(len(hierarchy1), len(hierarchy2))
        
        # 後ろから比較
        for i in range(1, min_len + 1):
            if hierarchy1[-i] == hierarchy2[-i]:
                score += i  # 深い階層ほど重要
            else:
                break
                
        return score
                
    def _get_next_sequence(self, parent_sequence: str, all_items: List[Dict]) -> str:
        """
        親sequenceの下で次に使用可能なsequence番号を計算
        
        Args:
            parent_sequence: 親のsequence（""の場合はルート）
            all_items: 既存アイテムリスト（互換性のため残す）
            
        Returns:
            次のsequence番号
        """
        # self.all_existing_itemsを優先的に使用
        items_to_check = getattr(self, 'all_existing_items', all_items)
        
        if parent_sequence:
            prefix = parent_sequence + "."
        else:
            prefix = ""
            
        # 同じ親を持つ既存のsequenceを収集
        sibling_sequences = []
        for item in items_to_check:
            seq = item.get("sequence") or ""  # Noneや空の場合は空文字列に
            if seq and seq.startswith(prefix):  # 空文字列もチェック
                # 直接の子供のみ（さらに深い階層は除外）
                remainder = seq[len(prefix):]
                if "." not in remainder and remainder.isdigit():
                    sibling_sequences.append(int(remainder))
                    
        # 次の番号を決定
        if sibling_sequences:
            next_num = max(sibling_sequences) + 1
        else:
            next_num = 1
            
        return f"{prefix}{next_num}" if prefix else str(next_num)
            
    def _read_new_requirement_description(self, ws, template_ref: str, req_type: str) -> Optional[str]:
        """
        新規要件用テンプレートからDescriptionを読み込み
        
        Args:
            ws: Description編集ワークシート
            template_ref: テンプレート参照（例: "#S1", "#U1"）
            req_type: "System" または "User"
            
        Returns:
            HTML形式のDescription
        """
        # プレフィックスとテンプレート番号を分離
        if req_type == "System":
            template_num = template_ref.strip("#S")
            search_pattern = f"【新規System Requirement #S{template_num}】"
        else:
            template_num = template_ref.strip("#U")
            search_pattern = f"【新規User Requirement #U{template_num}】"
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f"A{row}"].value
            if cell_value and search_pattern in str(cell_value):
                # テーブル開始行を特定（ヘッダーの3行下）
                table_start = row + 3
                
                if req_type == "System":
                    # 5行分のデータを読み込み
                    table_data = []
                    for i in range(5):
                        row_data = []
                        # 総列数まで読み込み
                        for col in range(1, self.system_desc_total_cols + 1):
                            value = ws.cell(row=table_start + i, column=col).value
                            row_data.append(str(value) if value is not None else "")
                        table_data.append(row_data)
                        
                    # HTMLテーブルに変換
                    return self._convert_to_html_table_system(table_data)
                else:
                    # 3行分のデータを読み込み
                    table_data = []
                    for i in range(3):
                        row_data = []
                        # 総列数まで読み込み
                        for col in range(1, self.user_desc_total_cols + 1):
                            value = ws.cell(row=table_start + i, column=col).value
                            row_data.append(str(value) if value is not None else "")
                        table_data.append(row_data)
                        
                    # HTMLテーブルに変換
                    return self._convert_to_html_table_user(table_data)
                
        logger.warning(f"テンプレート {template_ref} が見つかりません")
        return None
            
    def _read_description_from_sheet(self, ws, jama_id: str, req_type: str) -> Optional[str]:
        """
        Description編集シートから新しいDescriptionを読み込み（拡張性対応版）
        
        Args:
            ws: Description編集ワークシート
            jama_id: JAMA ID
            req_type: "System" または "User"
            
        Returns:
            HTML形式のDescription
        """
        # シート内で【JAMA_ID】のパターンを検索
        search_pattern = f"【{jama_id}】"
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f"A{row}"].value
            if cell_value and search_pattern in str(cell_value):
                # テーブル開始行を特定（【JAMA_ID】要件名の2行下）
                table_start = row + 2
                
                if req_type == "System":
                    # 5行分のデータを読み込み
                    table_data = []
                    for i in range(5):
                        row_data = []
                        # 総列数まで読み込み
                        for col in range(1, self.system_desc_total_cols + 1):
                            value = ws.cell(row=table_start + i, column=col).value
                            row_data.append(str(value) if value is not None else "")
                        table_data.append(row_data)
                        
                    # HTMLテーブルに変換
                    return self._convert_to_html_table_system(table_data)
                else:
                    # 3行分のデータを読み込み
                    table_data = []
                    for i in range(3):
                        row_data = []
                        # 総列数まで読み込み
                        for col in range(1, self.user_desc_total_cols + 1):
                            value = ws.cell(row=table_start + i, column=col).value
                            row_data.append(str(value) if value is not None else "")
                        table_data.append(row_data)
                        
                    # HTMLテーブルに変換
                    return self._convert_to_html_table_user(table_data)
                
        return None
        
    def _convert_to_html_table_system(self, table_data: List[List[str]]) -> str:
        """
        System Requirement用のテーブルデータをHTMLに変換（拡張性対応版）
        
        Args:
            table_data: テーブルデータ
            
        Returns:
            HTMLテーブル
        """
        # 列の位置を動的に計算
        pos = self._get_system_column_positions()
        
        html = "<table border='1' cellpadding='5' cellspacing='0'>\n"
        
        for row_idx, row in enumerate(table_data):
            html += "<tr>\n"
            
            # 特殊な結合処理
            if row_idx == 0:  # I/O Type行
                html += f"<td>{row[0]}</td>\n"
                # IN列に薄い青の背景色
                html += f"<td style='background-color: #E3F2FD;'>{row[1]}</td>\n"
                # OUT列に薄い緑の背景色（動的にcolspan計算）
                out_colspan = self.system_desc_total_cols - 2  # 総列数 - 最初の2列
                html += f"<td colspan='{out_colspan}' style='background-color: #E8F5E9;'>{row[2]}</td>\n"
            elif row_idx == 1:  # 項目名行
                html += f"<td>{row[0]}</td>\n"
                # (a)Trigger actionに薄い青の背景色
                html += f"<td style='background-color: #E3F2FD;'>{row[1]}</td>\n"
                # (b)Behavior of ego-vehicleに薄い緑の背景色（動的にcolspan計算）
                b_colspan = self.SYSTEM_DESC_COLS['b']
                html += f"<td colspan='{b_colspan}' style='background-color: #E8F5E9;'>{row[pos['b_start'] - 1]}</td>\n"
                # (c)HMIに薄い緑の背景色（動的にcolspan計算）
                c_colspan = self.SYSTEM_DESC_COLS['c']
                html += f"<td colspan='{c_colspan}' style='background-color: #E8F5E9;'>{row[pos['c_start'] - 1]}</td>\n"
                # (d)Otherに薄い緑の背景色（動的にcolspan計算）
                d_colspan = self.SYSTEM_DESC_COLS['d']
                html += f"<td colspan='{d_colspan}' style='background-color: #E8F5E9;'>{row[pos['d_start'] - 1]}</td>\n"
            else:  # データ行（色なし）
                for cell in row[:self.system_desc_total_cols]:  # 必要な列数のみ
                    html += f"<td>{cell}</td>\n"
                    
            html += "</tr>\n"
            
        html += "</table>"
        return html
        
    def _convert_to_html_table_user(self, table_data: List[List[str]]) -> str:
        """
        User Requirement用のテーブルデータをHTMLに変換（拡張性対応版）
        
        Args:
            table_data: テーブルデータ
            
        Returns:
            HTMLテーブル
        """
        # 列の位置を動的に計算
        pos = self._get_user_column_positions()
        
        html = "<table border='1' cellpadding='5' cellspacing='0'>\n"
        
        for row_idx, row in enumerate(table_data):
            html += "<tr>\n"
            
            # 特殊な結合処理
            if row_idx == 0:  # I/O Type行
                html += f"<td>{row[0]}</td>\n"
                # IN列に薄い青の背景色
                html += f"<td style='background-color: #E3F2FD;'>{row[1]}</td>\n"
                # OUT列に薄い緑の背景色（動的にcolspan計算）
                out_colspan = self.user_desc_total_cols - 2  # 総列数 - 最初の2列
                html += f"<td colspan='{out_colspan}' style='background-color: #E8F5E9;'>{row[2]}</td>\n"
            elif row_idx == 1:  # 項目名行
                html += f"<td>{row[0]}</td>\n"
                # (a)Trigger actionに薄い青の背景色
                html += f"<td style='background-color: #E3F2FD;'>{row[1]}</td>\n"
                # (c)HMIに薄い緑の背景色（動的にcolspan計算）
                c_colspan = self.USER_DESC_COLS['c']
                html += f"<td colspan='{c_colspan}' style='background-color: #E8F5E9;'>{row[pos['c_start'] - 1]}</td>\n"
            else:  # データ行（色なし）
                for cell in row[:self.user_desc_total_cols]:  # 必要な列数のみ
                    html += f"<td>{cell}</td>\n"
                    
            html += "</tr>\n"
            
        html += "</table>"
        return html
        
    def _convert_to_html_table(self, table_data: List[List[str]]) -> str:
        """
        テーブルデータをHTMLに変換（互換性のため残す）
        
        Args:
            table_data: テーブルデータ
            
        Returns:
            HTMLテーブル
        """
        # 5行の場合はSystem用、3行の場合はUser用として処理
        if len(table_data) == 5:
            return self._convert_to_html_table_system(table_data)
        elif len(table_data) == 3:
            return self._convert_to_html_table_user(table_data)
        else:
            # フォールバック：単純なテーブル変換
            html = "<table border='1' cellpadding='5' cellspacing='0'>\n"
            for row in table_data:
                html += "<tr>\n"
                for cell in row:
                    html += f"<td>{cell}</td>\n"
                html += "</tr>\n"
            html += "</table>"
            return html
        
    def validate_new_items(self, new_items: List[Dict]) -> List[str]:
        """
        新規作成アイテムのバリデーション（改良版）
        
        Args:
            new_items: 新規作成アイテムリスト
            
        Returns:
            エラーメッセージのリスト
        """
        errors = []
        
        # all_existing_itemsが存在しない場合は空リストとして扱う
        existing_items = getattr(self, 'all_existing_items', [])
        
        for item in new_items:
            # 親要件の存在チェック
            if item.get("parent_id"):
                parent_found = False
                parent_name = ""
                
                # 既存要件から検索
                for existing in existing_items:
                    if existing.get("jama_id") == item["parent_id"]:
                        parent_found = True
                        parent_name = existing.get("name", "Unknown")
                        break
                        
                if not parent_found:
                    # エラーメッセージに詳細情報を含める
                    error_msg = f"要件「{item.get('name', 'Unknown')}」の親要件（ID: {item['parent_id']}）が見つかりません"
                    if item.get('calculated_sequence'):
                        error_msg += f"\n  計算されたSequence: {item['calculated_sequence']}"
                    errors.append(error_msg)
                    
            # 必須フィールドのチェック
            if not item.get("name"):
                errors.append(f"要件名が設定されていません（Sequence: {item.get('calculated_sequence', 'Unknown')}）")
                
            # 階層レベルチェック（新規追加）
            if item.get("calculated_sequence"):
                # sequenceから階層レベルを計算
                dot_count = item["calculated_sequence"].count(".")
                hierarchy_level = dot_count + 1
                
                # 設定オブジェクトがある場合のみチェック
                if self.config and hierarchy_level < 10 and not self.config.is_item_type_defined_for_level(hierarchy_level):
                    # 1-9階層で未定義の場合のみエラー（10-11は動的判定なのでエラーにしない）
                    errors.append(f"階層{hierarchy_level}への新規作成はサポートされていません。item_type_idが未定義です。")
                
        return errors
    
    def _determine_item_type_by_name_for_excel(self, name: str, hierarchy_level: int) -> int:
        """
        要件名から10-11階層のアイテムタイプIDを判定（Excel用）
        
        Args:
            name: 要件名
            hierarchy_level: 階層レベル（10 or 11）
            
        Returns:
            アイテムタイプID
        """
        if not name or not self.config:
            return self.config.default_item_type_for_10_11 if self.config else 301
        
        # 大文字小文字を区別しない判定
        name_upper = name.upper()
        
        # より長い一致を優先するため、長い順にチェック
        if name_upper.startswith("SYFR"):
            return 301  # STD Oneteam System Requirement
        elif name_upper.startswith("SYSP"):
            return 301  # STD Oneteam System Requirement
        elif name_upper.startswith("FR"):
            return 266  # STD User Requirement
        elif name_upper.startswith("SP"):
            return 266  # STD User Requirement
        else:
            return self.config.default_item_type_for_10_11
