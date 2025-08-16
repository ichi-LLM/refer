"""
Trigger_editシートから異常系トリガー要件を解析
"""

import openpyxl
import logging
from typing import List, Dict, Optional
from pathlib import Path

logger = logging.getLogger(__name__)


class RequirementParser:
    """要件パーサークラス"""
    
    def parse_trigger_requirements(self, excel_file: str) -> List[Dict]:
        """
        ExcelファイルのTrigger_editシートからトリガー要件を解析
        
        Args:
            excel_file: Excelファイルパス
            
        Returns:
            トリガー要件のリスト
        """
        logger.info(f"トリガー要件の解析開始: {excel_file}")
        
        if not Path(excel_file).exists():
            raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_file}")
            
        try:
            # Excelファイルを開く
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            
            # Trigger_editシートを取得
            if "Trigger_edit" not in wb.sheetnames:
                logger.warning("Trigger_editシートが見つかりません")
                return []
                
            ws = wb["Trigger_edit"]
            
            # トリガー要件を解析
            triggers = self._parse_trigger_sheet(ws)
            
            wb.close()
            
            logger.info(f"トリガー要件を{len(triggers)}件解析しました")
            return triggers
            
        except Exception as e:
            logger.error(f"Excelファイルの解析に失敗: {str(e)}")
            raise
            
    def _parse_trigger_sheet(self, ws) -> List[Dict]:
        """
        Trigger_editシートを解析
        
        Args:
            ws: ワークシート
            
        Returns:
            トリガー要件リスト
        """
        triggers = []
        current_trigger = None
        current_row = 1
        max_row = ws.max_row
        
        # シート全体をスキャン
        while current_row <= max_row:
            # A列の値を確認
            cell_a = ws[f"A{current_row}"].value
            
            if cell_a and isinstance(cell_a, str):
                # 【JAMA_ID】で始まる行を検出
                if cell_a.startswith("【") and "】" in cell_a:
                    # 新しいトリガーブロックを検出
                    if current_trigger:
                        # 前のトリガーを保存
                        triggers.append(current_trigger)
                        
                    # JAMA_IDと要件名を抽出
                    parts = cell_a.split("】", 1)
                    jama_id = parts[0].replace("【", "").strip()
                    requirement_name = parts[1].strip() if len(parts) > 1 else ""
                    
                    logger.debug(f"トリガー検出: JAMA_ID={jama_id}, 名前={requirement_name}")
                    
                    # 新しいトリガー情報を初期化
                    current_trigger = {
                        "jama_id": jama_id,
                        "requirement_name": requirement_name,
                        "signal_name": "",
                        "value": "",
                        "action_type": "",
                        "remarks": ""
                    }
                    
                    # テーブルデータを読み込み（次の数行）
                    self._parse_trigger_table(ws, current_row + 2, current_trigger)
                    
                    # 次のブロックまでスキップ
                    current_row += 10  # おおよその間隔
                    continue
                    
            current_row += 1
            
        # 最後のトリガーを保存
        if current_trigger:
            triggers.append(current_trigger)
            
        return triggers
        
    def _parse_trigger_table(self, ws, start_row: int, trigger: Dict) -> None:
        """
        トリガーテーブルを解析（コンパクトテーブル形式）
        
        Args:
            ws: ワークシート
            start_row: テーブル開始行
            trigger: トリガー情報を格納する辞書
        """
        # テーブル形式:
        # | 項目 | 設定値 |
        # | 信号名 | CAN_BRAKE_ERROR |
        # | 値 | 1 |
        # | アクションタイプ | can_communication_error |
        # | 備考 | ... |
        
        for i in range(5):  # 最大5行をチェック
            row = start_row + i
            item_name = ws[f"A{row}"].value
            item_value = ws[f"B{row}"].value
            
            if not item_name:
                continue
                
            item_name = str(item_name).strip()
            item_value = str(item_value).strip() if item_value else ""
            
            if item_name == "信号名":
                trigger["signal_name"] = item_value
            elif item_name == "値":
                # 数値に変換を試みる
                try:
                    trigger["value"] = int(item_value)
                except ValueError:
                    try:
                        trigger["value"] = float(item_value)
                    except ValueError:
                        trigger["value"] = item_value
            elif item_name == "アクションタイプ":
                trigger["action_type"] = item_value
            elif item_name == "備考":
                trigger["remarks"] = item_value
                
        logger.debug(f"テーブル解析結果: {trigger}")
        
    def parse_requirement_hierarchy(self, excel_file: str, hierarchy_sequence: str) -> List[Dict]:
        """
        指定された階層のトリガー要件を抽出
        
        Args:
            excel_file: Excelファイルパス
            hierarchy_sequence: 階層のSequence（例: "1.2.3"）
            
        Returns:
            該当階層のトリガー要件リスト
        """
        logger.info(f"階層 {hierarchy_sequence} のトリガー要件を抽出")
        
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            
            # Requirement_of_Driverシートから階層情報を取得
            if "Requirement_of_Driver" not in wb.sheetnames:
                logger.error("Requirement_of_Driverシートが見つかりません")
                return []
                
            req_sheet = wb["Requirement_of_Driver"]
            
            # 指定階層に属する要件のJAMA_IDを収集
            trigger_jama_ids = []
            for row in range(2, req_sheet.max_row + 1):
                sequence = req_sheet[f"C{row}"].value
                if sequence and str(sequence).startswith(hierarchy_sequence):
                    jama_id = req_sheet[f"A{row}"].value
                    if jama_id:
                        trigger_jama_ids.append(str(jama_id))
                        
            wb.close()
            
            # Trigger_editシートから該当するトリガーを抽出
            all_triggers = self.parse_trigger_requirements(excel_file)
            filtered_triggers = [
                trigger for trigger in all_triggers 
                if trigger.get("jama_id") in trigger_jama_ids
            ]
            
            logger.info(f"階層 {hierarchy_sequence} から {len(filtered_triggers)} 件のトリガーを抽出")
            return filtered_triggers
            
        except Exception as e:
            logger.error(f"階層トリガーの抽出に失敗: {str(e)}")
            raise
            
    def validate_trigger(self, trigger: Dict) -> List[str]:
        """
        トリガー情報の妥当性を検証
        
        Args:
            trigger: トリガー情報
            
        Returns:
            エラーメッセージのリスト（空なら問題なし）
        """
        errors = []
        
        # 必須フィールドのチェック
        if not trigger.get("signal_name"):
            errors.append(f"JAMA_ID {trigger.get('jama_id', 'UNKNOWN')}: 信号名が未設定")
            
        if trigger.get("value") == "":
            errors.append(f"JAMA_ID {trigger.get('jama_id', 'UNKNOWN')}: 値が未設定")
            
        if not trigger.get("action_type"):
            errors.append(f"JAMA_ID {trigger.get('jama_id', 'UNKNOWN')}: アクションタイプが未設定")
            
        # アクションタイプの妥当性チェック
        valid_action_types = [
            "can_communication_error",
            "brake",
            "accelerator", 
            "steering",
            "sensor_fault",
            "powertrain_fault",
            "environmental_fault",
            "shift",
            "appcssw",
            "sw_turn_signal",
            "engine_startup_operation",
            "MM_display_touched_coord"
        ]
        
        action_type = trigger.get("action_type", "")
        if action_type and action_type not in valid_action_types:
            # 部分一致でも許可（例: can_brake_communication_error）
            if not any(valid in action_type for valid in valid_action_types):
                logger.warning(f"未知のアクションタイプ: {action_type}")
                
        return errors